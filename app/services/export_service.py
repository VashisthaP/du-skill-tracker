"""
SkillHive - Excel Export Service
==================================
Generates Excel files (.xlsx) for demands and applications data.
Uses openpyxl for Excel generation with formatted headers and auto-sized columns.
"""

import io
from datetime import datetime, timezone
from flask import send_file, current_app
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from app.models import Demand, Application


# =====================================================
# ACCENTURE-THEMED EXCEL STYLES
# =====================================================

# Header style: Accenture purple background with white text
HEADER_FONT = Font(name='Calibri', bold=True, color='FFFFFF', size=11)
HEADER_FILL = PatternFill(start_color='A100FF', end_color='A100FF', fill_type='solid')
HEADER_ALIGNMENT = Alignment(horizontal='center', vertical='center', wrap_text=True)

# Data cell style
DATA_FONT = Font(name='Calibri', size=10)
DATA_ALIGNMENT = Alignment(vertical='top', wrap_text=True)

# Border style
THIN_BORDER = Border(
    left=Side(style='thin'),
    right=Side(style='thin'),
    top=Side(style='thin'),
    bottom=Side(style='thin')
)

# Priority color fills
PRIORITY_FILLS = {
    'critical': PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'),
    'high': PatternFill(start_color='FFF0D0', end_color='FFF0D0', fill_type='solid'),
    'medium': PatternFill(start_color='E0F0FF', end_color='E0F0FF', fill_type='solid'),
    'low': PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid'),
}


# =====================================================
# EXPORT DEMANDS
# =====================================================

def export_demands_to_excel():
    """
    Export all demands to a formatted Excel file.
    Includes project details, skills, evaluator info, and application counts.

    Returns:
        Flask response with the Excel file as attachment.
    """
    demands = (
        Demand.query
        .order_by(Demand.created_at.desc())
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Demands"

    # Define column headers
    headers = [
        'ID', 'Project Name', 'Project Code', 'RRD',
        'Required Skills', 'Career Level', 'Positions', 'Start Date',
        'End Date', 'Priority', 'Status', 'Evaluator Name', 'Evaluator Email',
        'Evaluator Contact', 'Description', 'Applications', 'Created By',
        'Created At'
    ]

    # Write headers with styling
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, demand in enumerate(demands, 2):
        row_data = [
            demand.id,
            demand.project_name,
            demand.project_code or '',
            demand.rrd,
            demand.skills_display,
            f'CL{demand.career_level}',
            demand.num_positions,
            demand.start_date.strftime('%Y-%m-%d') if demand.start_date else '',
            demand.end_date.strftime('%Y-%m-%d') if demand.end_date else '',
            demand.priority.upper(),
            demand.status.replace('_', ' ').title(),
            demand.evaluator_name or '',
            demand.evaluator_email or '',
            demand.evaluator_contact or '',
            demand.description or '',
            demand.application_count,
            demand.creator.display_name if demand.creator else '',
            demand.created_at.strftime('%Y-%m-%d %H:%M') if demand.created_at else '',
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # Apply priority color to the priority column (column 10)
        priority_cell = ws.cell(row=row_idx, column=10)
        if demand.priority in PRIORITY_FILLS:
            priority_cell.fill = PRIORITY_FILLS[demand.priority]

    # Auto-adjust column widths
    _auto_adjust_columns(ws)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Generate the file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    filename = f'SkillHive_Demands_{timestamp}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# =====================================================
# EXPORT APPLICATIONS
# =====================================================

def export_applications_to_excel(demand_id=None):
    """
    Export applications to a formatted Excel file.
    Optionally filtered by demand_id.

    Args:
        demand_id: If provided, only export applications for this demand.

    Returns:
        Flask response with the Excel file as attachment.
    """
    query = Application.query.join(Demand)

    if demand_id:
        query = query.filter(Application.demand_id == demand_id)

    applications = query.order_by(Application.applied_at.desc()).all()

    wb = Workbook()
    ws = wb.active
    ws.title = "Applications"

    # Define column headers
    headers = [
        'ID', 'Applicant Name', 'Enterprise ID', 'Project Name',
        'RRD', 'Career Level', 'Current Project', 'Years of Experience',
        'Skills', 'Resume', 'Status', 'Remarks', 'Applied At', 'Last Updated'
    ]

    # Write headers with styling
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = HEADER_FONT
        cell.fill = HEADER_FILL
        cell.alignment = HEADER_ALIGNMENT
        cell.border = THIN_BORDER

    # Write data rows
    for row_idx, app in enumerate(applications, 2):
        row_data = [
            app.id,
            app.applicant_name,
            app.enterprise_id or '',
            app.demand.project_name if app.demand else '',
            app.demand.rrd if app.demand else '',
            f'CL{app.demand.career_level}' if app.demand else '',
            app.current_project or '',
            app.years_of_experience or 0,
            app.skills_text or '',
            app.resume_filename or 'Not attached',
            app.status.replace('_', ' ').title(),
            app.remarks or '',
            app.applied_at.strftime('%Y-%m-%d %H:%M') if app.applied_at else '',
            app.updated_at.strftime('%Y-%m-%d %H:%M') if app.updated_at else '',
        ]

        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.font = DATA_FONT
            cell.alignment = DATA_ALIGNMENT
            cell.border = THIN_BORDER

        # Color-code status column (column 11)
        status_cell = ws.cell(row=row_idx, column=11)
        status_fills = {
            'applied': PatternFill(start_color='E0F0FF', end_color='E0F0FF', fill_type='solid'),
            'under_evaluation': PatternFill(start_color='FFF0D0', end_color='FFF0D0', fill_type='solid'),
            'selected': PatternFill(start_color='E0FFE0', end_color='E0FFE0', fill_type='solid'),
            'rejected': PatternFill(start_color='FFE0E0', end_color='FFE0E0', fill_type='solid'),
        }
        if app.status in status_fills:
            status_cell.fill = status_fills[app.status]

    # Auto-adjust column widths
    _auto_adjust_columns(ws)

    # Freeze the header row
    ws.freeze_panes = 'A2'

    # Generate the file
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')
    suffix = f'_demand_{demand_id}' if demand_id else ''
    filename = f'SkillHive_Applications{suffix}_{timestamp}.xlsx'

    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename
    )


# =====================================================
# HELPER FUNCTIONS
# =====================================================

def _auto_adjust_columns(ws, max_width=50):
    """
    Auto-adjust column widths based on cell content.
    Caps maximum width to prevent extremely wide columns.
    """
    for col in ws.columns:
        max_length = 0
        col_letter = col[0].column_letter

        for cell in col:
            try:
                cell_length = len(str(cell.value or ''))
                max_length = max(max_length, cell_length)
            except Exception:
                pass

        # Add padding and cap the width
        adjusted_width = min(max_length + 3, max_width)
        ws.column_dimensions[col_letter].width = adjusted_width

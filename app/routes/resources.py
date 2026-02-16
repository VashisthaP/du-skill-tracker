"""
SkillHive - Resource Routes
=============================
Bulk resource upload, listing, evaluation, and export.
PMO uploads Excel resource lists against demands (RRDs).
Evaluators review resources and provide feedback (select / reject).
"""

from flask import (
    Blueprint, render_template, redirect, url_for, flash,
    request, current_app, send_file
)
from flask_login import login_required, current_user
from datetime import datetime, timezone
from io import BytesIO
import openpyxl
from openpyxl.styles import Font, PatternFill

from app import db
from app.models import Demand, Resource
from app.forms import ResourceUploadForm, ResourceEvaluationForm, ProjectForm
from app.utils.decorators import pmo_required

resources_bp = Blueprint('resources', __name__, template_folder='templates')

# =====================================================
# EXCEL HEADER → MODEL FIELD MAPPING
# =====================================================
# Supports various header spellings / formats
HEADER_MAP = {
    'PERSONNEL_NO': 'personnel_no',
    'PRE HIRE ID': 'personnel_no',
    'PERSONNEL_NO/PRE HIRE ID': 'personnel_no',
    'PERSONNEL NO': 'personnel_no',
    'NAME': 'name',
    'EMPLOYEE_PRIMARY_SKILL': 'primary_skill',
    'EMPLOYEE PRIMARY SKILL': 'primary_skill',
    'PRIMARY SKILL': 'primary_skill',
    'MANAGEMENT LEVEL': 'management_level',
    'MANAGEMENT_LEVEL': 'management_level',
    'HOME_LOC': 'home_location',
    'HOME LOC': 'home_location',
    'HOME LOCATION': 'home_location',
    'CURRENT_LOCK_STATUS': 'lock_status',
    'CURRENT LOCK STATUS': 'lock_status',
    'LOCK STATUS': 'lock_status',
    'ROLL_OFF_DATE': 'availability_status',
    'ROLL OFF DATE': 'availability_status',
    'ROLLOFF DATE': 'availability_status',
    'E_MAIL_ADDRESS': 'email',
    'E MAIL ADDRESS': 'email',
    'EMAIL': 'email',
    'EMAIL ADDRESS': 'email',
    'CONTACT_DETAILS': 'contact_details',
    'CONTACT DETAILS': 'contact_details',
    'CONTACT': 'contact_details',
    'PHONE': 'contact_details',
    'JOINING DATE': 'joining_date',
    'JOINING DATE (BENCH/JOINERS)': 'joining_date',
    'JOINING_DATE': 'joining_date',
}


def _match_header(header_text):
    """Match an Excel header to a model field using the HEADER_MAP."""
    normalized = header_text.strip().upper()
    # Exact match first
    if normalized in HEADER_MAP:
        return HEADER_MAP[normalized]
    # Partial / contains match
    for key, field in HEADER_MAP.items():
        if key in normalized or normalized in key:
            return field
    return None


# =====================================================
# SELECT PROJECT (PMO only) - Landing page for Upload Resources
# =====================================================
@resources_bp.route('/')
@login_required
@pmo_required
def select_project():
    """
    Display list of projects for PMO to select before uploading resources.
    Shows all projects with option to create a new one.
    """
    page = request.args.get('page', 1, type=int)
    per_page = 10

    # Get all projects (demands), newest first
    query = Demand.query.order_by(Demand.created_at.desc())

    # Search filter
    search = request.args.get('search', '').strip()
    if search:
        search_pattern = f'%{search}%'
        query = query.filter(
            db.or_(
                Demand.project_name.ilike(search_pattern),
                Demand.du_name.ilike(search_pattern),
                Demand.client_name.ilike(search_pattern),
                Demand.manager_name.ilike(search_pattern),
            )
        )

    projects = query.paginate(page=page, per_page=per_page, error_out=False)

    return render_template('resources/select_project.html',
                           projects=projects,
                           search=search)


# =====================================================
# CREATE PROJECT (PMO only)
# =====================================================
@resources_bp.route('/create-project', methods=['GET', 'POST'])
@login_required
@pmo_required
def create_project():
    """
    Create a new project with simplified fields.
    After creation, redirect to upload resources for this project.
    """
    form = ProjectForm()

    if form.validate_on_submit():
        try:
            # Create demand/project with simplified fields
            project = Demand(
                project_name=form.project_name.data,
                du_name=form.du_name.data,
                client_name=form.client_name.data,
                manager_name=form.manager_name.data,
                description=form.description.data,
                # Set defaults for required fields
                rrd=form.project_name.data,  # Use project name as RRD
                career_level='11',  # Default to SSE
                num_positions=1,
                priority='medium',
                status='open',
                created_by=current_user.id,
            )

            db.session.add(project)
            db.session.commit()

            flash(f'Project "{project.project_name}" created successfully! Now upload resources.', 'success')
            return redirect(url_for('resources.upload', demand_id=project.id))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f"Error creating project: {e}")
            flash('An error occurred while creating the project. Please try again.', 'danger')

    return render_template('resources/create_project.html', form=form)


# =====================================================
# UPLOAD RESOURCES (PMO only)
# =====================================================
@resources_bp.route('/upload/<int:demand_id>', methods=['GET', 'POST'])
@login_required
@pmo_required
def upload(demand_id):
    """Upload an Excel file with resources for a specific demand (RRD)."""
    demand = Demand.query.get_or_404(demand_id)
    form = ResourceUploadForm()

    if form.validate_on_submit():
        file = form.excel_file.data
        try:
            wb = openpyxl.load_workbook(file, read_only=True, data_only=True)
            ws = wb.active

            # --- Read & map headers from row 1 ---
            header_row = next(ws.iter_rows(min_row=1, max_row=1))
            field_map = {}  # column_index → model_field
            for idx, cell in enumerate(header_row):
                if cell.value:
                    field = _match_header(str(cell.value))
                    if field:
                        field_map[idx] = field

            if not field_map:
                flash('Could not detect column headers. '
                      'Please ensure the first row contains headers like NAME, '
                      'EMPLOYEE_PRIMARY_SKILL, E_MAIL_ADDRESS, etc.', 'danger')
                return render_template('resources/upload.html',
                                       form=form, demand=demand)

            # --- Parse data rows ---
            count = 0
            errors = 0
            for row in ws.iter_rows(min_row=2):
                row_data = {}
                for idx, cell in enumerate(row):
                    if idx in field_map and cell.value is not None:
                        value = str(cell.value).strip()
                        if value:
                            row_data[field_map[idx]] = value

                # Skip empty rows or rows without a name
                if not row_data or not row_data.get('name'):
                    continue

                try:
                    resource = Resource(
                        demand_id=demand.id,
                        personnel_no=row_data.get('personnel_no', ''),
                        name=row_data.get('name', ''),
                        primary_skill=row_data.get('primary_skill', ''),
                        management_level=row_data.get('management_level', ''),
                        home_location=row_data.get('home_location', ''),
                        lock_status=row_data.get('lock_status', ''),
                        availability_status=row_data.get('availability_status', ''),
                        email=row_data.get('email', ''),
                        contact_details=row_data.get('contact_details', ''),
                        joining_date=row_data.get('joining_date', ''),
                        uploaded_by=current_user.id,
                    )
                    db.session.add(resource)
                    count += 1
                except Exception as e:
                    errors += 1
                    current_app.logger.warning(f'Error parsing resource row: {e}')

            db.session.commit()
            wb.close()

            msg = f'Successfully uploaded {count} resource(s) for {demand.rrd}.'
            if errors:
                msg += f' ({errors} row(s) had errors and were skipped.)'
            flash(msg, 'success')
            return redirect(url_for('resources.list_resources',
                                     demand_id=demand.id))

        except Exception as e:
            db.session.rollback()
            current_app.logger.error(f'Excel upload error: {e}')
            flash(f'Error processing Excel file: {str(e)}', 'danger')

    return render_template('resources/upload.html', form=form, demand=demand)


# =====================================================
# LIST RESOURCES (PMO / Evaluator / Admin)
# =====================================================
@resources_bp.route('/demand/<int:demand_id>')
@login_required
def list_resources(demand_id):
    """List all resources uploaded for a specific demand."""
    demand = Demand.query.get_or_404(demand_id)

    # Optional status filter
    status_filter = request.args.get('status', '')
    query = Resource.query.filter_by(demand_id=demand.id)
    if status_filter:
        query = query.filter_by(evaluation_status=status_filter)

    resources = query.order_by(Resource.uploaded_at.desc()).all()
    eval_form = ResourceEvaluationForm()

    # Stats for filter badges
    stats = {
        'total': Resource.query.filter_by(demand_id=demand.id).count(),
        'pending': Resource.query.filter_by(
            demand_id=demand.id, evaluation_status='pending').count(),
        'under_evaluation': Resource.query.filter_by(
            demand_id=demand.id, evaluation_status='under_evaluation').count(),
        'accepted': Resource.query.filter_by(
            demand_id=demand.id, evaluation_status='accepted').count(),
        'rejected': Resource.query.filter_by(
            demand_id=demand.id, evaluation_status='rejected').count(),
        'skill_mismatch': Resource.query.filter_by(
            demand_id=demand.id, evaluation_status='skill_mismatch').count(),
        'unavailable': Resource.query.filter_by(
            demand_id=demand.id, evaluation_status='unavailable').count(),
        'already_locked': Resource.query.filter_by(
            demand_id=demand.id, evaluation_status='already_locked').count(),
    }

    return render_template(
        'resources/list.html',
        demand=demand,
        resources=resources,
        eval_form=eval_form,
        stats=stats,
        status_filter=status_filter,
    )


# =====================================================
# EVALUATE RESOURCE (Evaluator / PMO / Admin)
# =====================================================
@resources_bp.route('/<int:resource_id>/evaluate', methods=['POST'])
@login_required
def evaluate(resource_id):
    """Evaluator provides feedback on a resource."""
    if not (current_user.is_evaluator or current_user.is_pmo
            or current_user.is_admin):
        flash('You do not have permission to evaluate resources.', 'danger')
        return redirect(url_for('main.dashboard'))

    resource = Resource.query.get_or_404(resource_id)
    form = ResourceEvaluationForm()

    if form.validate_on_submit():
        resource.evaluation_status = form.evaluation_status.data
        resource.evaluation_remarks = form.evaluation_remarks.data
        resource.evaluated_by = current_user.id
        resource.evaluated_at = datetime.now(timezone.utc)
        db.session.commit()

        flash(f'Evaluation updated for {resource.name}: '
              f'{resource.status_display}', 'success')
    else:
        flash('Invalid form submission.', 'danger')

    return redirect(url_for('resources.list_resources',
                             demand_id=resource.demand_id))


# =====================================================
# DELETE SINGLE RESOURCE (PMO only)
# =====================================================
@resources_bp.route('/<int:resource_id>/delete', methods=['POST'])
@login_required
@pmo_required
def delete(resource_id):
    """Remove a single resource from the list."""
    resource = Resource.query.get_or_404(resource_id)
    demand_id = resource.demand_id
    name = resource.name
    db.session.delete(resource)
    db.session.commit()
    flash(f'Resource "{name}" removed.', 'info')
    return redirect(url_for('resources.list_resources', demand_id=demand_id))


# =====================================================
# DELETE ALL RESOURCES FOR A DEMAND (PMO only)
# =====================================================
@resources_bp.route('/delete-all/<int:demand_id>', methods=['POST'])
@login_required
@pmo_required
def delete_all(demand_id):
    """Clear all resources for a demand."""
    demand = Demand.query.get_or_404(demand_id)
    count = Resource.query.filter_by(demand_id=demand.id).delete()
    db.session.commit()
    flash(f'Removed {count} resource(s) from {demand.rrd}.', 'info')
    return redirect(url_for('demands.detail', demand_id=demand.id))


# =====================================================
# EXPORT RESOURCES TO EXCEL
# =====================================================
@resources_bp.route('/export/<int:demand_id>')
@login_required
def export_excel(demand_id):
    """Export resources for a demand to an Excel file."""
    if not (current_user.is_evaluator or current_user.is_pmo
            or current_user.is_admin):
        flash('Access denied.', 'danger')
        return redirect(url_for('main.dashboard'))

    demand = Demand.query.get_or_404(demand_id)
    resources = (Resource.query
                 .filter_by(demand_id=demand.id)
                 .order_by(Resource.name)
                 .all())

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = f'Resources - {demand.rrd}'

    # Headers
    headers = [
        'Personnel No', 'Name', 'Primary Skill', 'Management Level',
        'Home Location', 'Lock Status', 'Availability', 'Email',
        'Contact Details', 'Joining Date', 'Evaluation Status',
        'Evaluation Remarks', 'Evaluated By',
    ]
    ws.append(headers)

    # Style header row
    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='A100FF', end_color='A100FF',
                              fill_type='solid')
    for cell in ws[1]:
        cell.font = header_font
        cell.fill = header_fill

    # Data rows
    for r in resources:
        ws.append([
            r.personnel_no, r.name, r.primary_skill, r.management_level,
            r.home_location, r.lock_status, r.availability_status, r.email,
            r.contact_details, r.joining_date, r.status_display,
            r.evaluation_remarks or '',
            r.evaluator.display_name if r.evaluator else '',
        ])

    # Auto-width columns
    for col in ws.columns:
        max_len = max(len(str(cell.value or '')) for cell in col) + 2
        ws.column_dimensions[col[0].column_letter].width = min(max_len, 40)

    output = BytesIO()
    wb.save(output)
    output.seek(0)

    filename = (f'resources_{demand.rrd}_'
                f'{datetime.now().strftime("%Y%m%d")}.xlsx')
    return send_file(
        output,
        mimetype='application/vnd.openxmlformats-officedocument.'
                 'spreadsheetml.sheet',
        as_attachment=True,
        download_name=filename,
    )
